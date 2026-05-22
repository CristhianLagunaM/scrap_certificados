from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def write_report(dataframe: pd.DataFrame, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    dataframe.to_excel(path, index=False)
    apply_report_style(path)
    return path


def apply_report_style(path: Path) -> None:
    workbook = load_workbook(path)
    worksheet = workbook.active

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    error_fill = PatternFill("solid", fgColor="F4CCCC")
    omitted_fill = PatternFill("solid", fgColor="FFF2CC")
    success_fill = PatternFill("solid", fgColor="D9EAD3")
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    estado_col = None
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        if cell.value == "Estado procesamiento":
            estado_col = cell.column

    for row in worksheet.iter_rows(min_row=2):
        estado = row[estado_col - 1].value if estado_col else ""
        fill = None
        if estado == "No descargada":
            fill = error_fill
        elif estado == "Omitida":
            fill = omitted_fill
        elif estado == "Descargada":
            fill = success_fill

        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
            if fill:
                cell.fill = fill

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    workbook.save(path)

